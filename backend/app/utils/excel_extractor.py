"""
Excel Data Extraction Utility
Extracts data from Excel files (.xlsx, .xls) using openpyxl and pandas
"""
import openpyxl
import pandas as pd
from typing import Dict, List
import io


def extract_data_from_excel(file_content: bytes, file_extension: str = '.xlsx') -> Dict[str, any]:
    """
    Extract data from Excel file
    
    Args:
        file_content: Excel file content as bytes
        file_extension: File extension (.xlsx or .xls)
        
    Returns:
        Dictionary with extracted data and metadata
    """
    result = {
        "text": "",
        "sheets": [],
        "total_sheets": 0,
        "success": True,
        "error": None
    }
    
    try:
        # Read Excel file with pandas
        excel_file = io.BytesIO(file_content)
        
        # Read all sheets
        excel_data = pd.read_excel(excel_file, sheet_name=None, engine='openpyxl')
        
        result["total_sheets"] = len(excel_data)
        
        for sheet_name, df in excel_data.items():
            # Convert dataframe to text
            sheet_text = f"\\n\\n=== Sheet: {sheet_name} ===\\n\\n"
            
            # Add column headers
            sheet_text += " | ".join(str(col) for col in df.columns) + "\\n"
            sheet_text += "-" * 80 + "\\n"
            
            # Add rows
            for idx, row in df.iterrows():
                row_text = " | ".join(str(val) for val in row.values)
                sheet_text += row_text + "\\n"
            
            result["sheets"].append({
                "name": sheet_name,
                "row_count": len(df),
                "column_count": len(df.columns),
                "columns": list(df.columns),
                "text": sheet_text,
                "char_count": len(sheet_text)
            })
            
            result["text"] += sheet_text + "\\n\\n"
        
        # Statistics
        result["text"] = result["text"].strip()
        result["char_count"] = len(result["text"])
        result["word_count"] = len(result["text"].split())
        
    except Exception as e:
        result["success"] = False
        result["error"] = f"Excel extraction failed: {str(e)}"
    
    return result


def extract_specific_sheets(file_content: bytes, sheet_names: List[str]) -> Dict[str, pd.DataFrame]:
    """
    Extract specific sheets from Excel file
    
    Args:
        file_content: Excel file content as bytes
        sheet_names: List of sheet names to extract
        
    Returns:
        Dictionary mapping sheet names to DataFrames
    """
    try:
        excel_file = io.BytesIO(file_content)
        
        # Read specific sheets
        sheets_data = {}
        for sheet_name in sheet_names:
            try:
                df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl')
                sheets_data[sheet_name] = df
            except Exception as e:
                print(f"Error reading sheet '{sheet_name}': {str(e)}")
        
        return sheets_data
    
    except Exception as e:
        print(f"Error extracting Excel sheets: {str(e)}")
        return {}


def get_excel_summary(file_content: bytes) -> Dict[str, any]:
    """
    Get summary information about Excel file
    
    Args:
        file_content: Excel file content as bytes
        
    Returns:
        Dictionary with summary information
    """
    try:
        excel_file = io.BytesIO(file_content)
        wb = openpyxl.load_workbook(excel_file, read_only=True)
        
        summary = {
            "sheet_names": wb.sheetnames,
            "total_sheets": len(wb.sheetnames),
            "sheets_info": []
        }
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            summary["sheets_info"].append({
                "name": sheet_name,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column
            })
        
        wb.close()
        return summary
    
    except Exception as e:
        return {"error": f"Failed to get Excel summary: {str(e)}"}
